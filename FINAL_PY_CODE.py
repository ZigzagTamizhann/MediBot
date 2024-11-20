import os
import cv2
import math
import time
import socket
import numpy as np
import pandas as pd
import mediapipe as mp
from deepface import DeepFace
from datetime import datetime
from comtypes import CLSCTX_ALL
from ctypes import cast, POINTER
from openpyxl import Workbook, load_workbook
from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume

esp32_ip = '192.168.1.108'
esp32_port = 80

call = ''
Emotion = []
heart_rate_readings = []
temperature_readings = []

# Paths
images_base_path = "D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\images"
excel_path_emo = "D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\MEDIBOT_FINAL.xlsx"
volume_data_path = "D:\\Python Files\\MEDIBOT_FINAL\\Volume\\Image"
excel_path_vol = "D:\\Python Files\\MEDIBOT_FINAL\\Volume\\VolumeData.xlsx"

# Create the base images directory if it doesn't exist
if not os.path.exists(images_base_path):
    os.makedirs(images_base_path)
    
# Create directories if they don't exist
if not os.path.exists(volume_data_path):
    os.makedirs(volume_data_path)
    
# Load face cascade classifier
face_cascade = cv2.CascadeClassifier("D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\haarcascade_frontalface_default.xml")

def send_commandddd(a):
    try:
        print(f"Sending command: {a}")
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.connect((esp32_ip, esp32_port))
            s.sendall(a.encode())
            
            # Wait for a response (assuming the ESP32 sends back '1' or '0')
            response = s.recv(1024).decode().strip()
            print(f"Received response: {response}")
            return response  # Return the response for further processing
    except socket.timeout:
        print("Connection timed out. Retrying...")
        return None  # Handle timeout by returning None or appropriate value
    except Exception as e:
        print(f"Error - {e}")
        return None  # Handle other exceptions
    
def send_command(command):
    if command in ['11', 'Temp', 'Heart', 'Call', '00']:
        try:
            print("Sending : ", command)
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.connect((esp32_ip, esp32_port))
                s.sendall(command.encode())
                if command == 'Temp':
                    receive_data(s, "Temperature detection finished.", temperature_readings)
                elif command == 'Heart':
                    receive_data(s, "Heart rate detection finished.", heart_rate_readings)
                elif command == 'Call':
                    receive_data(s, "SMS sent and call made.", call)
                else:
                    response = s.recv(1024).decode().strip()
                    print(f"ESP32 response: {response}")
        except socket.timeout:
            print(f"Connection timed out. Retrying...")
        except Exception as e:
            print(f"Error - {e}")
        time.sleep(2)  # Wait before retrying
    else:
        if command == 'Emotion':
            zz = detect_emotions()
        elif command == 'Val':
            zz = volmain()
        return zz

def receive_data(sock, end_signal, data_list):
    try:
        sock.settimeout(65)
        start_time = time.time()
        while True:
            response = sock.recv(1024).decode().strip()
            if response:
                if response.lower() == end_signal.lower():
                    print(f"{end_signal} received.")
                    break
                elif ":" in response:
                    print(f"Received: {response}")
                    z = response.index(':')
                    if end_signal == 'Temperature detection finished.':
                        temperature_readings.append(response[z+1:].strip())
                    elif end_signal == 'Heart rate detection finished.':
                        heart_rate_readings.append(response)
                    elif end_signal == 'SMS sent and call made.':
                        call += response
                else:
                    print(f"ESP32 response: {response}")

            # Break after 65 seconds
            if time.time() - start_time > 120:
                print(f"{end_signal} detection timed out.")
                break
        if end_signal == 'Temperature detection finished.':
            print(temperature_readings)
        elif end_signal == 'Heart rate detection finished.':
            print(heart_rate_readings)
        elif end_signal == 'SMS sent and call made.':
            print(call)
    except socket.timeout:
        print(f"{end_signal} detection timed out.")

# Function for flow-based emotion detection
def detect_emotions():
    cap = cv2.VideoCapture(0)
    detected_emotions = []  # List to store detected emotions
    emotion_count = {}  # Dictionary to store counts of each emotion
    start_time = time.time()
    detection_interval = 5  # Perform detection every 5 seconds
    timeout_duration = 60  # Run for 60 seconds (1 minute)

    while time.time() - start_time < timeout_duration:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture frame. Skipping...")
            continue

        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        for (x, y, w, h) in faces:
            face_roi = frame[y:y + h, x:x + w]

            try:
                analysis = DeepFace.analyze(face_roi, actions=['emotion'], enforce_detection=False)
                emotion = analysis[0]['dominant_emotion']
                detected_emotions.append(emotion)
                emotion_count[emotion] = emotion_count.get(emotion, 0) + 1

                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, f"{emotion}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
            except Exception as e:
                print(f"Error analyzing face: {e}")

        cv2.imshow('Real-time Emotion Detection', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

        if time.time() - start_time >= timeout_duration:
            break

    cap.release()
    cv2.destroyAllWindows()

    print("\nDetected Emotions List:", detected_emotions)
    save_emotions_to_excel(detected_emotions)
    return detected_emotions  # Return detected emotions for further use

def save_emotions_to_excel(detected_emotions):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    day_str = now.strftime("%A")

    for emotion in detected_emotions:
        data_to_save = {'Time': time_str, 'Date': date_str, 'Day': day_str, 'Emotion': emotion}

        try:
            if os.path.exists(excel_path_emo):
                book = load_workbook(excel_path_emo)
                if date_str in book.sheetnames:
                    sheet = book[date_str]
                else:
                    sheet = book.create_sheet(title=date_str)
                    sheet.append(['Time', 'Date', 'Day', 'Emotion'])
            else:
                book = Workbook()
                sheet = book.active
                sheet.title = date_str
                sheet.append(['Time', 'Date', 'Day', 'Emotion'])

            sheet.append([data_to_save['Time'], data_to_save['Date'], data_to_save['Day'], data_to_save['Emotion']])
            book.save(excel_path_emo)

        except PermissionError as e:
            print(f"Permission Error: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")

def volmain():
    # Solution APIs
    mp_drawing = mp.solutions.drawing_utils
    mp_hands = mp.solutions.hands

    # Volume Control Setup using PyCaw
    devices = AudioUtilities.GetSpeakers()
    interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
    volume = cast(interface, POINTER(IAudioEndpointVolume))
    volRange = volume.GetVolumeRange()
    minVol, maxVol = volRange[0], volRange[1]
    volBar, volPer = 400, 0
    volume_list = []  # List to store detected volume levels
    start_time = time.time()  # Time tracker for the duration of the program

    cam = cv2.VideoCapture(0)

    if not cam.isOpened():
        print("Error: Could not open video stream from IP camera")
        return

    # Mediapipe Hand Landmark Model
    with mp_hands.Hands(
        model_complexity=0,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5) as hands:

        frame_count = 0

        while cam.isOpened():
            success, image = cam.read()
            if not success:
                break

            # Update the current time for each frame
            current_time = time.time()

            # Convert the image to RGB for MediaPipe processing
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            results = hands.process(image)
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

            # Process detected hand landmarks
            if results.multi_hand_landmarks:
                for hand_landmarks in results.multi_hand_landmarks:
                    mp_drawing.draw_landmarks(
                        image,
                        hand_landmarks,
                        mp_hands.HAND_CONNECTIONS,
                        mp_drawing.DrawingSpec(color=(121, 22, 76), thickness=2, circle_radius=4),
                        mp_drawing.DrawingSpec(color=(250, 44, 250), thickness=2, circle_radius=2)
                    )

                    # Extract the required landmarks for volume control
                    landmarks = hand_landmarks.landmark
                    x1, y1 = landmarks[4].x * image.shape[1], landmarks[4].y * image.shape[0]
                    x2, y2 = landmarks[8].x * image.shape[1], landmarks[8].y * image.shape[0]

                    # Draw a line between the two landmarks
                    cv2.line(image, (int(x1), int(y1)), (int(x2), int(y2)), (255, 0, 0), 3)

                    # Calculate the distance between the thumb and index finger
                    length = math.hypot(x2 - x1, y2 - y1)

                    # Interpolate volume based on the distance
                    vol = np.interp(length, [20, 200], [minVol, maxVol])
                    volBar = np.interp(length, [20, 200], [400, 150])
                    volPer = np.interp(length, [20, 200], [0, 100])

                    # Set the volume
                    volume.SetMasterVolumeLevel(vol, None)
                    volume_list.append(volPer)  # Store volume level for later use

                    # Draw volume bar
                    cv2.rectangle(image, (50, 150), (85, 400), (0, 255, 0), 2)
                    cv2.rectangle(image, (50, int(volBar)), (85, 400), (0, 255, 0), cv2.FILLED)
                    cv2.putText(image, f"{int(volPer)} %", (40, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 0), 3)

            cv2.imshow('Volume Control', image)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

            # Break after 60 seconds
            if current_time - start_time > 60:
                break

        cam.release()
        cv2.destroyAllWindows()

    # Save the volume levels to Excel
    save_volume_to_excel(volume_list)
    return volume_list  # Return the volume levels for further use

def save_volume_to_excel(volume_data):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    day_str = now.strftime("%A")

    try:
        if os.path.exists(excel_path_vol):
            book = load_workbook(excel_path_vol)
            if date_str in book.sheetnames:
                sheet = book[date_str]
            else:
                sheet = book.create_sheet(title=date_str)
                sheet.append(['Time', 'Date', 'Day', 'Volume Level'])
        else:
            book = Workbook()
            sheet = book.active
            sheet.title = date_str
            sheet.append(['Time', 'Date', 'Day', 'Volume Level'])

        for volume_level in volume_data:
            sheet.append([time_str, date_str, day_str, volume_level])

        book.save(excel_path_vol)

    except PermissionError as e:
        print(f"Permission Error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def main():
    print("Checking connection with ESP32...")
    for i in ["S_1", "F_2", "S_3", "T_4", "F_5", "F_6", "S_7", "S_8", "E_9", "N_10", "T_E"]:
        print("Task : ",i)
        yy=send_commandddd(i)
        if yy=='1':
            for j in ['11', 'temp', 'heart', 'call', 'emotion','val', '00']:
                zz=send_command(j.capitalize())
                print(zz)
        
        time.sleep(5)

if __name__ == "__main__":
    main()
