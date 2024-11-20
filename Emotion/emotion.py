import cv2
from deepface import DeepFace
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time

# Paths
images_base_path = "D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\images"
excel_path = "D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\MEDIBOT_FINAL.xlsx"
url = 0  # Use webcam

# Create the base images directory if it doesn't exist
if not os.path.exists(images_base_path):
    os.makedirs(images_base_path)

# Load face cascade classifier
face_cascade = cv2.CascadeClassifier("D:\\Python Files\\MEDIBOT_FINAL\\Emotion\\haarcascade_frontalface_default.xml")

# Function to log data to the Excel sheet
def log_to_excel(emotion, date_str, time_str):
    # Check if the Excel file exists
    if not os.path.exists(excel_path):
        # Create a new workbook if it does not exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Emotion"])  # Add header row
        workbook.save(excel_path)
    
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    # Add the new row of emotion data
    sheet.append([date_str, time_str, emotion])
    
    # Save the updated workbook
    workbook.save(excel_path)

# Function for flow-based emotion detection
def detect_emotions():
    # Start capturing video
    cap = cv2.VideoCapture(url)

    if not cap.isOpened():
        print("Error: Could not open video stream.")
        return

    # Initialize variables
    detected_emotions = []  # List to store detected emotions
    emotion_count = {}  # Dictionary to store counts of each emotion

    # Start timing
    start_time = time.time()
    detection_interval = 5  # Perform detection every 5 seconds

    while time.time() - start_time < 60:  # Run for 60 seconds (1 minute)
        # Capture frame-by-frame
        ret, frame = cap.read()

        if not ret:
            print("Failed to capture frame. Skipping...")
            break  # Skip to the next iteration if frame capture fails

        # Resize frame for consistent processing
        detection_frame_size = (1280, 720)
        frame = cv2.resize(frame, detection_frame_size)

        # Convert frame to grayscale
        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Detect faces in the frame
        faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        emotion_detected = False  # Flag to check if any emotion is detected
        for (x, y, w, h) in faces:
            # Extract the face ROI (Region of Interest)
            face_roi = frame[y:y + h, x:x + w]

            # Analyze the face for emotion
            try:
                analysis = DeepFace.analyze(face_roi, actions=['emotion'], enforce_detection=False)
                emotion = analysis[0]['dominant_emotion']
                detected_emotions.append(emotion)  # Add detected emotion to the list
                emotion_detected = True  # Set the flag to True

                # Update emotion count
                if emotion in emotion_count:
                    emotion_count[emotion] += 1
                else:
                    emotion_count[emotion] = 1

                # Draw rectangle around face and label with predicted emotion
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, f"{emotion}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
            except Exception as e:
                print(f"Error analyzing face: {e}")

        # Display the resulting frame
        cv2.imshow('Real-time Emotion Detection', frame)

        # Check if any emotion was detected
        if emotion_detected:
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H-%M-%S")

            # Save the detected frame in the daily folder
            daily_images_path = os.path.join(images_base_path, date_str)
            if not os.path.exists(daily_images_path):
                os.makedirs(daily_images_path)

            if frame is not None and frame.size > 0:
                frame_filename = os.path.join(daily_images_path, f"{time_str}_{emotion}.jpg")
                cv2.imwrite(frame_filename, frame)

            # Log emotion to the Excel file
            log_to_excel(emotion, date_str, time_str)

        # Break if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the capture and close all windows
    cap.release()
    cv2.destroyAllWindows()

# Example usage
if __name__ == "__main__":
    detect_emotions()
